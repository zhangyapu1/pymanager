import warnings
# 抑制所有 openpyxl 相关的 UserWarning
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.*')

import os
import re
import time
import tkinter as tk
from tkinter import filedialog, Tk, simpledialog, messagebox, ttk


def _showinfo(title, msg, **kw):
    print(f"[{title}] {msg}")
    messagebox.showinfo(title, msg, **kw)


def _showerror(title, msg, **kw):
    print(f"[{title}] {msg}")
    messagebox.showerror(title, msg, **kw)


def _showwarning(title, msg, **kw):
    print(f"[{title}] {msg}")
    messagebox.showwarning(title, msg, **kw)


def _askyesno(title, msg, **kw):
    print(f"[{title}] {msg}")
    return messagebox.askyesno(title, msg, **kw)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.exceptions import InvalidFileException

def progress_bar(current, total, bar_length=50):
    """显示进度条"""
    if total == 0:
        percent = 1.0
    else:
        percent = current / total
    filled_length = int(bar_length * percent)
    bar = '█' * filled_length + '-' * (bar_length - filled_length)
    # 添加 flush 确保在部分 IDE 或终端中实时显示
    print(f'\r[{bar}] {current}/{total} ({percent:.1%})', end='', flush=True)

# 常量定义，便于维护
REFRESH_INTERVAL_ROW = 10  # 行进度条刷新间隔
REFRESH_INTERVAL_COL_SCAN = 5  # 列扫描时的刷新间隔
DEFAULT_OUTPUT_NAME = "Excel汇总结果"


class ProgressDialog:
    """
    改进版进度对话框：
    - 双层进度条（文件级 + 行级）
    - 实时更新当前文件、工作表、已处理行数
    - 支持取消
    """

    def __init__(self, parent, total_files):
        self.parent = parent
        self.total_files = total_files
        self.cancelled = False

        self.window = tk.Toplevel(parent)
        self.window.title("正在汇总...")
        self.window.geometry("560x300")
        self.window.resizable(False, False)
        self.window.transient(parent)
        self.window.grab_set()
        self.window.protocol("WM_DELETE_WINDOW", self.on_cancel)

        pad = {"padx": 20, "pady": 3, "anchor": "w"}

        # 标题
        tk.Label(self.window, text="正在处理 Excel 文件，请稍候…",
                 font=("微软雅黑", 10, "bold")).pack(padx=20, pady=(16, 4))

        # 当前文件
        self.var_file = tk.StringVar(value="–")
        tk.Label(self.window, textvariable=self.var_file,
                 font=("微软雅黑", 9), fg="#1a56db").pack(**pad)

        # 当前工作表
        self.var_sheet = tk.StringVar(value="")
        tk.Label(self.window, textvariable=self.var_sheet,
                 font=("微软雅黑", 9), fg="#0e7d5a").pack(**pad)

        # 文件进度条
        tk.Label(self.window, text="文件进度：",
                 font=("微软雅黑", 9), fg="gray").pack(padx=20, anchor="w")
        self.bar_file = ttk.Progressbar(self.window, length=520,
                                         mode="determinate",
                                         maximum=max(total_files, 1))
        self.bar_file.pack(padx=20, pady=2)

        # 行级进度条（当前工作表内）
        tk.Label(self.window, text="当前工作表行进度：",
                 font=("微软雅黑", 9), fg="gray").pack(padx=20, anchor="w")
        self.bar_row = ttk.Progressbar(self.window, length=520,
                                        mode="determinate", maximum=100)
        self.bar_row.pack(padx=20, pady=2)

        # 状态文字
        self.var_status = tk.StringVar(value="")
        tk.Label(self.window, textvariable=self.var_status,
                 font=("微软雅黑", 9), fg="gray").pack(padx=20, pady=(4, 0), anchor="w")

        # 取消按钮
        tk.Button(self.window, text="取消", width=10,
                  command=self.on_cancel).pack(pady=10)

        # 居中显示
        self.window.update_idletasks()
        w, h = self.window.winfo_width(), self.window.winfo_height()
        x = (self.window.winfo_screenwidth() - w) // 2
        y = (self.window.winfo_screenheight() - h) // 2
        self.window.geometry(f"{w}x{h}+{x}+{y}")
        self.window.update()

    # ── 外部调用接口 ────────────────────────────────────────────────

    def set_file(self, file_path, file_index):
        """切换到新文件时调用"""
        name = os.path.basename(file_path)
        text = f"文件 {file_index}/{self.total_files}：{name}"
        # 兼容不同的 StringVar 实现，虽然 tk.StringVar 通常只有 set
        if hasattr(self.var_file, 'config'):
            self.var_file.config(text=text)
        else:
            self.var_file.set(text)
            
        self.bar_file["value"] = file_index - 1   # 开始处理时先移到前一格
        self.bar_row["value"] = 0
        self._flush()

    def set_sheet(self, sheet_name, total_rows_in_range):
        """切换到新工作表时调用"""
        self.var_sheet.set(f"工作表：{sheet_name}  （目标行范围 {total_rows_in_range} 行）")
        self.bar_row["maximum"] = max(total_rows_in_range, 1)
        self.bar_row["value"] = 0
        self._flush()

    def update_row(self, rows_done_in_sheet, total_rows_extracted):
        """每处理一行（或若干行）后调用"""
        self.bar_row["value"] = rows_done_in_sheet
        self.var_status.set(f"累计已提取有效数据：{total_rows_extracted} 行")
        # 每 REFRESH_INTERVAL_ROW 行才真正刷新一次界面，减少 tkinter 调度开销
        if rows_done_in_sheet % REFRESH_INTERVAL_ROW == 0:
            self._flush()

    def finish_file(self, file_index):
        """一个文件处理完毕时调用"""
        self.bar_file["value"] = file_index
        self._flush()

    def close(self):
        if self.window and self.window.winfo_exists():
            self.window.destroy()

    def on_cancel(self):
        self.cancelled = True
        self.close()

    # ── 内部工具 ────────────────────────────────────────────────────

    def _flush(self):
        if self.window and self.window.winfo_exists():
            try:
                self.window.update()
            except tk.TclError:
                # 窗口可能在更新过程中被关闭
                pass


# ════════════════════════════════════════════════════════════════════


def 汇总Excel文件_按条件():
    """汇总多个 Excel 文件中符合条件的工作表指定行范围（带双层实时进度条）"""

    root = Tk()
    root.withdraw()

    _showinfo(
        "使用说明",
        "本工具将汇总多个 Excel 文件中符合条件的工作表\n\n"
        "1. 输入工作表关键词（支持模糊匹配）\n"
        "2. 输入起始行和结束行\n"
        "3. 选择多个 Excel 文件\n"
        "4. 自动生成汇总结果（带实时双层进度条）",
    )

    sheet_keyword = simpledialog.askstring(
        "输入工作表关键词",
        "请输入工作表名称关键词（支持模糊匹配）：\n"
        "例如：销售、Sheet、2024\n\n留空表示汇总所有工作表",
        initialvalue="",
    )
    if sheet_keyword is None:
        return

    start_row = simpledialog.askinteger(
        "输入起始行",
        "请输入要汇总的起始行号：",
        initialvalue=48, minvalue=1, maxvalue=1048576,
    )
    if start_row is None:
        return

    end_row = simpledialog.askinteger(
        "输入结束行",
        "请输入要汇总的结束行号：",
        initialvalue=54, minvalue=start_row, maxvalue=1048576,
    )
    if end_row is None:
        return

    include_header = _askyesno(
        "是否包含表头",
        "是否在汇总结果中包含原始数据的表头行？\n\n"
        "【是】：复制原始数据第 1 行作为表头\n"
        "【否】：使用默认表头（来源文件、工作表名称等）",
    )

    output_name = simpledialog.askstring(
        "输出文件名", "请输入输出文件名（不含扩展名）：",
        initialvalue=DEFAULT_OUTPUT_NAME,
    )
    if not output_name:
        output_name = DEFAULT_OUTPUT_NAME

    files = filedialog.askopenfilenames(
        title="请选择要汇总的 Excel 文件（可多选）",
        filetypes=[
            ("Excel 文件", "*.xlsx *.xlsm *.xls"),
            ("所有文件", "*.*"),
        ],
    )
    if not files:
        _showwarning("提示", "未选择任何文件，程序退出")
        return

    print(f"\n{'='*60}")
    total_files = len(files)
    print(f"开始处理，共选择 {total_files} 个文件")
    print(f"筛选条件：工作表名称包含【{sheet_keyword or '所有'}】")
    print(f"行范围：第 {start_row} 行 — 第 {end_row} 行")
    print(f"{'='*60}\n")
    
    start_time = time.time()

    # ── 样式常量 ────────────────────────────────────────────────────
    header_font  = Font(bold=True, size=11, color="FFFFFF")
    header_fill  = PatternFill(start_color="4472C4", end_color="4472C4",
                               fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font    = Font(size=10)
    data_align   = Alignment(horizontal="left", vertical="center",
                              wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── 结果工作簿 ──────────────────────────────────────────────────
    wb_result = Workbook()
    ws_result  = wb_result.active
    ws_result.title = "汇总结果"

    default_headers = ["来源文件", "工作表名称", "原始行号"]
    header_copied   = False

    if not include_header:
        ws_result.append(default_headers)
        # 优化：直接对第一行单元格应用样式
        for cell in ws_result[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border
        header_copied = True

    current_result_row = 2 if header_copied else 1
    total_rows         = 0
    processed_sheets   = []
    error_files        = []

    progress = ProgressDialog(root, len(files))

    # ── 逐文件处理 ──────────────────────────────────────────────────
    for file_index, file_path in enumerate(files, 1):
        # 更新进度条
        progress_bar(file_index, total_files)
        
        file_name = os.path.basename(file_path)
        print(f"\n[{file_index}/{total_files}] 正在处理: {file_name}")

        progress.set_file(file_path, file_index)

        if progress.cancelled:
            _showinfo("已取消", "用户终止了汇总操作。")
            break

        wb_src = None
        try:
            # 严格检查文件扩展名，避免加载非 Excel 文件
            lower_name = file_name.lower()
            if not (lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm") or lower_name.endswith(".xls")):
                raise InvalidFileException("文件格式不支持")

            if lower_name.endswith(".xls") and not (lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm")):
                print(f"  跳过旧格式文件（请转换为 .xlsx）: {file_name}")
                error_files.append(f"{file_name} (旧格式，不支持)")
                progress.finish_file(file_index)
                continue

            wb_src = load_workbook(file_path, data_only=True, read_only=True)
            sheet_found = False

            for sheet_name in wb_src.sheetnames:
                if sheet_keyword and sheet_keyword.lower() not in sheet_name.lower():
                    continue

                sheet_found = True
                print(f"  匹配工作表: [{sheet_name}]")

                ws_src         = wb_src[sheet_name]
                range_size     = max(end_row - start_row + 1, 1)
                # 处理 max_row 可能为 None 的情况
                src_max_row = ws_src.max_row
                actual_end_row = min(end_row, src_max_row if src_max_row else end_row)
                
                # 如果起始行都超过了最大行，跳过
                if start_row > (src_max_row if src_max_row else 0):
                    print(f"    └─ 起始行超出范围，跳过（文件最大行数: {src_max_row if src_max_row else 0}，指定起始行: {start_row}）")
                    continue

                progress.set_sheet(sheet_name, range_size)

                # ── 第一遍：确定最大有效列 ──────────────────────────
                max_col = 1
                # 使用 values_only=True 提高速度
                for r_idx, row in enumerate(
                    ws_src.iter_rows(
                        min_row=start_row, max_row=actual_end_row,
                        values_only=True
                    )
                ):
                    # 优化：使用 enumerate 并提前打断如果不需要更宽的列
                    # 注意：row 可能是 None 填充的，或者长度不一
                    for c_idx, val in enumerate(row, 1):
                        if val is not None and str(val).strip():
                            if c_idx > max_col:
                                max_col = c_idx
                    
                    # 每 REFRESH_INTERVAL_COL_SCAN 行刷新一次
                    if r_idx % REFRESH_INTERVAL_COL_SCAN == 0:
                        progress.update_row(r_idx, total_rows)
                        if progress.cancelled:
                            break

                if progress.cancelled:
                    break
                
                # 如果没有找到任何有数据的列，重置为默认最小列数，避免后续错误
                if max_col < len(default_headers):
                     max_col = len(default_headers)

                # ── 复制原始表头（可选）──────────────────────────────
                if include_header and not header_copied and start_row > 1:
                    header_row_data = []
                    # 读取第一行作为表头
                    # 注意：read_only 模式下，访问 cell 可能需要小心，但 iter_rows 或 cell 方法通常可用
                    # 为了保险，我们只读取需要的列
                    for col in range(1, max_col + 1):
                        val = ws_src.cell(1, col).value
                        header_row_data.append(
                            val if val is not None else f"列{col}"
                        )
                    ws_result.append(default_headers + header_row_data)
                    for cell in ws_result[1]:
                        cell.font      = header_font
                        cell.fill      = header_fill
                        cell.alignment = header_align
                        cell.border    = thin_border
                    header_copied      = True
                    current_result_row = 2
                    progress._flush()

                # ── 第二遍：提取数据（实时更新进度）─────────────────
                row_count = 0
                for r_offset, row_vals in enumerate(
                    ws_src.iter_rows(
                        min_row=start_row, max_row=actual_end_row,
                        max_col=max_col, values_only=True
                    )
                ):
                    actual_row_num = start_row + r_offset

                    # 优化：列表推导式比循环 append 快
                    row_data = [
                        (v if v is not None else "") for v in row_vals
                    ]
                    
                    # 补齐列数（read_only 模式可能返回较短的行，如果该行后面全是空）
                    # 注意：iter_rows(max_col=...) 应该保证长度，但为了安全
                    curr_len = len(row_data)
                    if curr_len < max_col:
                        row_data.extend([""] * (max_col - curr_len))

                    # 检查是否有有效数据
                    has_data = False
                    for v in row_data:
                        if str(v).strip():
                            has_data = True
                            break
                    
                    if has_data:
                        ws_result.append(
                            [file_name, sheet_name, actual_row_num] + row_data
                        )
                        current_result_row += 1
                        total_rows         += 1
                        row_count          += 1

                    # 实时刷新：每行都更新进度条数值，每 REFRESH_INTERVAL_ROW 行才真正绘制
                    progress.update_row(r_offset + 1, total_rows)
                    if progress.cancelled:
                        break

                if row_count:
                    print(f"    └─ 提取了 {row_count} 行数据")
                    processed_sheets.append(
                        f"{file_name} - {sheet_name} ({row_count} 行)"
                    )
                else:
                    print(f"    └─ 指定范围内无有效数据")

                if progress.cancelled:
                    break

            if not sheet_found:
                print(f"  无匹配工作表（关键词: {sheet_keyword or '全部'}）")
                error_files.append(f"{file_name} (无匹配工作表)")

        except InvalidFileException:
             print(f"  处理失败: 文件格式无效或损坏")
             error_files.append(f"{file_name} (格式无效)")
        except Exception as exc:
            print(f"  处理失败: {exc}")
            error_files.append(f"{file_name} ({str(exc)[:50]})") # 截断过长错误信息
        finally:
            # 确保关闭工作簿，释放文件句柄
            if wb_src:
                try:
                    wb_src.close()
                except:
                    pass

        progress.finish_file(file_index)
        print()

        if progress.cancelled:
            _showinfo("已取消", "用户终止了汇总操作。")
            break

    # ── 关闭进度对话框 ──────────────────────────────────────────────
    progress.close()

    cancelled = progress.cancelled

    # ── 后处理与保存 ────────────────────────────────────────────────
    if not cancelled and total_rows > 0:

        # 自动调整列宽
        # 优化：只遍历有数据的列
        for column in ws_result.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    val_str = str(cell.value)
                    raw_len = len(val_str)
                    # 简单估算中文宽度
                    chinese_cnt = len(re.findall(r"[\u4e00-\u9fff]", val_str))
                    weighted_len = raw_len + chinese_cnt 
                    if weighted_len > max_len:
                        max_len = weighted_len
            
            # 设置列宽，限制最小和最大值
            new_width = min(max(max_len + 2, 8), 50)
            ws_result.column_dimensions[col_letter].width = new_width

        # 数据行样式
        # 优化：批量处理样式，避免过多的属性查找
        # 注意：openpyxl 没有直接的“区域样式”应用，必须逐个单元格
        # 但我们可以通过减少对象访问来微调，或者接受这个开销，因为这是保存前的最后一步
        for row in ws_result.iter_rows(min_row=2,
                                        max_row=current_result_row - 1):
            for cell in row:
                cell.font      = data_font
                cell.alignment = data_align
                cell.border    = thin_border

        # 行高
        ws_result.row_dimensions[1].height = 25
        # 优化：批量设置行高在某些版本可能不支持，这里保持循环但范围明确
        for r in range(2, current_result_row):
            ws_result.row_dimensions[r].height = 18

        # 冻结首行 & 自动筛选
        ws_result.freeze_panes    = "A2"
        ws_result.auto_filter.ref = ws_result.dimensions

        # 输出路径（避免覆盖）
        try:
            # 尝试使用源文件目录
            output_dir = os.path.dirname(files[0])
            # 确保输出目录存在
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # 生成输出文件路径
            output_path = os.path.join(output_dir, f"{output_name}.xlsx")
            counter = 1
            while os.path.exists(output_path):
                output_path = os.path.join(output_dir,
                                            f"{output_name}_{counter}.xlsx")
                counter += 1

            # 保存文件
            wb_result.save(output_path)
        except Exception as save_error:
            print(f"保存文件失败: {save_error}")
            # 尝试使用桌面目录作为备用
            try:
                desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop_dir):
                    desktop_dir = os.path.expanduser("~")  # 如果桌面目录不存在，使用用户主目录
                
                # 生成备用输出路径
                output_path = os.path.join(desktop_dir, f"{output_name}.xlsx")
                counter = 1
                while os.path.exists(output_path):
                    output_path = os.path.join(desktop_dir,
                                                f"{output_name}_{counter}.xlsx")
                    counter += 1
                
                # 保存文件到备用位置
                wb_result.save(output_path)
                print(f"[INFO] 保存到备用位置: {output_path}")
                _showinfo("保存成功", f"由于权限限制，文件已保存到桌面:\n{output_path}")
            except Exception as backup_error:
                print(f"备用位置保存失败: {backup_error}")
                _showerror("保存失败", f"无法保存输出文件:\n{save_error}\n\n备用位置也无法保存:\n{backup_error}\n\n请手动选择一个可写的输出目录。")
                return

        # 完成进度条
        progress_bar(total_files, total_files)
        print()  # 换行
        
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        print(f"\n{'='*60}")
        print(f"[OK] 汇总完成！")
        print(f"[DIR] 输出文件: {output_path}")
        print(f"[FILES] 处理文件数: {len(files)}")
        print(f"[SHEETS] 匹配工作表数: {len(processed_sheets)}")
        print(f"[ROWS] 汇总数据行数: {total_rows}")
        print(f"[TIME] 耗时: {elapsed_time:.2f}秒")
        if error_files:
            print(f"\n[WARN] 处理失败 ({len(error_files)} 个):")
            for e in error_files:
                print(f"   - {e}")
        if processed_sheets and len(processed_sheets) <= 20:
            print(f"\n[LIST] 详细列表:")
            for s in processed_sheets:
                print(f"   - {s}")
        elif processed_sheets:
            print(f"\n[LIST] 共处理 {len(processed_sheets)} 个工作表")
        print(f"{'='*60}")

        _showinfo(
            "完成",
            f"汇总完成！\n\n"
            f"输出文件：{os.path.basename(output_path)}\n"
            f"处理文件：{len(files)} 个\n"
            f"汇总数据：{total_rows} 行\n\n"
            f"文件保存在：{output_dir}",
        )

    elif not cancelled and total_rows == 0:
        _showwarning(
            "警告",
            "未提取到任何有效数据！\n\n"
            "请检查：\n"
            "1. 工作表关键词是否正确\n"
            "2. 指定行范围内是否有数据\n"
            "3. Excel 文件是否为空",
        )
    else:
        print("已取消操作，未保存任何文件。")
    
    # 清理 root 窗口
    try:
        root.destroy()
    except:
        pass


if __name__ == "__main__":
    汇总Excel文件_按条件()