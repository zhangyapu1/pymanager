import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl.utils import get_column_letter

def sanitize_sheet_name(name, max_len=31):
    """清理工作表名，避免 Excel 非法字符和超长"""
    illegal_chars = r'[\\/*?:\[\]]'
    sanitized = re.sub(illegal_chars, '_', name)
    if len(sanitized) > max_len:
        sanitized = sanitized[:max_len]
    return sanitized

def auto_adjust_column_width(writer, sheet_name, df):
    """
    自动调整列宽，增加异常保护，防止因空数据或异常值导致程序中断
    """
    try:
        workbook = writer.book
        worksheet = workbook[sheet_name]
        
        # 如果 DataFrame 为空，设置默认列宽后返回
        if df.empty:
            for i, col in enumerate(df.columns):
                col_letter = get_column_letter(i + 1)
                worksheet.column_dimensions[col_letter].width = 10
            return
        
        for i, col in enumerate(df.columns):
            col_letter = get_column_letter(i + 1)
            # 计算表头长度
            header_len = len(str(col))
            # 计算该列数据的最大字符长度（处理 NaN 和各类类型）
            try:
                # 将列转换为字符串，填充 NaN 为空字符串
                col_str = df[col].fillna('').astype(str)
                max_data_len = col_str.map(len).max() if not col_str.empty else 0
            except Exception:
                max_data_len = 0
            max_len = max(header_len, max_data_len)
            # 列宽 = 最大长度 + 2，限制最小5，最大50
            adjusted_width = min(max(max_len + 2, 5), 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
    except Exception as e:
        # 如果调整列宽失败，只打印警告，不影响整体合并
        print(f"  警告：工作表 {sheet_name} 调整列宽时出错：{e}")

def merge_excel_files(file_paths, output_file):
    """将多个 Excel 文件合并到一个工作簿，自动调整列宽"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for file_path in file_paths:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            try:
                sheets = pd.read_excel(file_path, sheet_name=None)
            except Exception as e:
                print(f"读取文件失败：{file_path}，错误：{e}")
                continue
            
            if len(sheets) == 1:
                sheet_name, df = list(sheets.items())[0]
                new_name = sanitize_sheet_name(base_name)
                df.to_excel(writer, sheet_name=new_name, index=False)
                auto_adjust_column_width(writer, new_name, df)
                print(f"已添加：{file_path} -> 工作表：{new_name}")
            else:
                for original_name, df in sheets.items():
                    new_name = sanitize_sheet_name(f"{base_name}_{original_name}")
                    df.to_excel(writer, sheet_name=new_name, index=False)
                    auto_adjust_column_width(writer, new_name, df)
                    print(f"已添加：{file_path} (原表：{original_name}) -> 工作表：{new_name}")
        print("所有文件处理完毕。")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="请选择要合并的 Excel 文件",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    if not file_paths:
        print("未选择任何文件，程序退出。")
        exit()
    
    output_file = filedialog.asksaveasfilename(
        title="保存合并后的文件",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not output_file:
        print("未指定输出文件，程序退出。")
        exit()
    
    merge_excel_files(list(file_paths), output_file)
    print(f"\n合并完成！结果保存在：{output_file}")
