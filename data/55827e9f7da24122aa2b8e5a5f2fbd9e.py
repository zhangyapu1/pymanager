import os
import re
from tkinter import filedialog, Tk, simpledialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def 汇总Excel文件_按条件():
    """汇总多个Excel文件中符合条件的工作表指定行范围"""
    
    # 隐藏tkinter主窗口
    root = Tk()
    root.withdraw()
    
    # 设置窗口图标（可选）
    try:
        root.iconbitmap(default='')
    except:
        pass
    
    # 1. 输入参数
    messagebox.showinfo("使用说明", 
                        "本工具将汇总多个Excel文件中符合条件的工作表\n\n"
                        "1. 输入工作表关键词（支持模糊匹配）\n"
                        "2. 输入起始行和结束行\n"
                        "3. 选择多个Excel文件\n"
                        "4. 自动生成汇总结果")
    
    # 工作表关键词
    sheet_keyword = simpledialog.askstring("输入工作表关键词", 
                                           "请输入工作表名称关键词（支持模糊匹配）：\n"
                                           "例如：销售、Sheet、*所有*、2024\n\n"
                                           "留空表示汇总所有工作表",
                                           initialvalue="")
    if sheet_keyword is None:
        print("操作已取消")
        return
    
    # 起始行
    start_row = simpledialog.askinteger("输入起始行", 
                                        "请输入要汇总的起始行号：\n"
                                        "例如：48",
                                        initialvalue=48,
                                        minvalue=1,
                                        maxvalue=1048576)
    if start_row is None:
        print("操作已取消")
        return
    
    # 结束行
    end_row = simpledialog.askinteger("输入结束行", 
                                      "请输入要汇总的结束行号：\n"
                                      "例如：54",
                                      initialvalue=54,
                                      minvalue=start_row,
                                      maxvalue=1048576)
    if end_row is None:
        print("操作已取消")
        return
    
    # 是否包含表头
    include_header = messagebox.askyesno("是否包含表头", 
                                         "是否在汇总结果中包含原始数据的表头行？\n\n"
                                         "选择【是】：会复制原始数据的第1行作为表头\n"
                                         "选择【否】：使用默认表头（来源文件、工作表名称等）")
    
    # 输出文件名
    output_name = simpledialog.askstring("输入输出文件名", 
                                         "请输入输出文件名（不含扩展名）：\n"
                                         "例如：汇总结果",
                                         initialvalue="Excel汇总结果")
    if not output_name:
        output_name = "Excel汇总结果"
    
    # 2. 选择多个Excel文件
    files = filedialog.askopenfilenames(
        title="请选择要汇总的Excel文件（可多选）",
        filetypes=[("Excel文件", "*.xlsx *.xlsm *.xls"), 
                   ("Excel 2007+", "*.xlsx"),
                   ("Excel 启用宏", "*.xlsm"),
                   ("Excel 97-2003", "*.xls"),
                   ("所有文件", "*.*")]
    )
    
    if not files:
        messagebox.showwarning("提示", "未选择任何文件，程序退出")
        print("未选择任何文件，程序退出")
        return
    
    print(f"\n{'='*60}")
    print(f"开始处理，共选择 {len(files)} 个文件")
    print(f"筛选条件：工作表名称包含【{sheet_keyword if sheet_keyword else '所有'}】")
    print(f"行范围：第 {start_row} 行 到 第 {end_row} 行")
    print(f"{'='*60}\n")
    
    # 3. 创建汇总工作簿
    wb_result = Workbook()
    ws_result = wb_result.active
    ws_result.title = "汇总结果"
    
    # 设置默认表头
    if include_header:
        # 先不设置表头，等遇到第一个有效数据时复制其表头
        header_copied = False
        default_headers = ["来源文件", "工作表名称", "原始行号"]
    else:
        default_headers = ["来源文件", "工作表名称", "原始行号"]
        ws_result.append(default_headers)
        header_copied = True
    
    # 设置表头样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用样式到表头
    if not include_header or header_copied:
        for cell in ws_result[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    
    current_row = 2 if header_copied else 1
    total_rows = 0
    processed_sheets = []
    error_files = []
    
    # 4. 处理每个文件
    for file_index, file_path in enumerate(files, 1):
        file_name = os.path.basename(file_path)
        print(f"[{file_index}/{len(files)}] 正在处理: {file_name}")
        
        try:
            # 判断文件格式，如果是.xls需要使用xlrd，这里主要处理.xlsx和.xlsm
            if file_name.lower().endswith('.xls') and not file_name.lower().endswith(('.xlsx', '.xlsm')):
                print(f"  ⚠ 跳过旧格式文件（请转换为.xlsx格式）: {file_name}")
                error_files.append(f"{file_name} (旧格式，不支持)")
                continue
            
            # 加载工作簿（只读模式，提高性能）
            wb_src = load_workbook(file_path, data_only=True, read_only=True)
            
            sheet_found = False
            
            for sheet_name in wb_src.sheetnames:
                # 判断工作表名称是否符合条件
                if sheet_keyword:
                    # 模糊匹配（不区分大小写）
                    if sheet_keyword.lower() not in sheet_name.lower():
                        continue
                
                sheet_found = True
                print(f"  ✓ 匹配工作表: [{sheet_name}]")
                
                ws_src = wb_src[sheet_name]
                
                # 获取该工作表的实际最大列数（在指定行范围内）
                max_col = 1
                for row in range(start_row, min(end_row, ws_src.max_row) + 1):
                    for col in range(1, ws_src.max_column + 1):
                        try:
                            cell_value = ws_src.cell(row, col).value
                            if cell_value is not None and str(cell_value).strip() != "":
                                if col > max_col:
                                    max_col = col
                        except:
                            pass
                
                # 复制表头（如果需要）
                if include_header and not header_copied and start_row > 1:
                    # 尝试获取原始数据的表头（第1行）
                    header_row_data = []
                    for col in range(1, max_col + 1):
                        header_value = ws_src.cell(1, col).value
                        header_row_data.append(header_value if header_value is not None else f"列{col}")
                    
                    # 添加来源信息列的表头
                    full_headers = default_headers + header_row_data
                    ws_result.append(full_headers)
                    
                    # 应用样式
                    for cell in ws_result[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border
                    
                    header_copied = True
                    current_row = 2
                
                # 提取数据
                row_count = 0
                for row in range(start_row, min(end_row, ws_src.max_row) + 1):
                    # 检查该行是否为空
                    is_empty = True
                    row_data = []
                    
                    for col in range(1, max_col + 1):
                        try:
                            cell_value = ws_src.cell(row, col).value
                            if cell_value is not None and str(cell_value).strip() != "":
                                is_empty = False
                            row_data.append(cell_value if cell_value is not None else "")
                        except:
                            row_data.append("")
                    
                    if not is_empty:
                        # 添加来源信息
                        row_with_source = [file_name, sheet_name, row] + row_data
                        ws_result.append(row_with_source)
                        current_row += 1
                        total_rows += 1
                        row_count += 1
                
                if row_count > 0:
                    print(f"    └─ 提取了 {row_count} 行数据")
                    processed_sheets.append(f"{file_name} - {sheet_name} ({row_count}行)")
                else:
                    print(f"    └─ 指定范围内无有效数据")
            
            if not sheet_found:
                print(f"  ✗ 未找到符合条件的工作表（关键词: {sheet_keyword}）")
                error_files.append(f"{file_name} (无匹配工作表)")
            
            wb_src.close()
            
        except Exception as e:
            print(f"  ✗ 处理失败: {str(e)}")
            error_files.append(f"{file_name} ({str(e)})")
            continue
        
        print()  # 空行分隔
    
    # 5. 调整格式
    if total_rows > 0:
        # 自动调整列宽
        for column in ws_result.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        # 计算单元格内容长度
                        cell_length = len(str(cell.value))
                        # 中文字符按2个长度计算
                        chinese_count = len(re.findall(r'[\u4e00-\u9fff]', str(cell.value)))
                        cell_length = cell_length + chinese_count
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            # 设置列宽（最小8，最大50）
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws_result.column_dimensions[column_letter].width = adjusted_width
        
        # 设置数据行样式
        data_font = Font(size=10)
        data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for row in ws_result.iter_rows(min_row=2, max_row=current_row-1):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
        
        # 设置行高
        ws_result.row_dimensions[1].height = 25
        for row in range(2, current_row):
            ws_result.row_dimensions[row].height = 18
        
        # 冻结首行
        ws_result.freeze_panes = 'A2'
        
        # 添加筛选功能
        ws_result.auto_filter.ref = ws_result.dimensions
        
        # 6. 保存结果
        output_dir = os.path.dirname(files[0])
        output_path = os.path.join(output_dir, f"{output_name}.xlsx")
        
        # 如果文件已存在，添加序号
        counter = 1
        while os.path.exists(output_path):
            output_path = os.path.join(output_dir, f"{output_name}_{counter}.xlsx")
            counter += 1
        
        wb_result.save(output_path)
        
        # 7. 显示汇总信息
        print(f"\n{'='*60}")
        print(f"✅ 汇总完成！")
        print(f"{'='*60}")
        print(f"📁 输出文件: {output_path}")
        print(f"📊 处理文件数: {len(files)}")
        print(f"📋 匹配工作表数: {len(processed_sheets)}")
        print(f"📝 汇总数据行数: {total_rows}")
        
        if error_files:
            print(f"\n⚠ 处理失败的文件 ({len(error_files)}个):")
            for err in error_files:
                print(f"   - {err}")
        
        # 显示详细的工作表列表
        if processed_sheets and len(processed_sheets) <= 20:
            print(f"\n📑 详细列表:")
            for sheet in processed_sheets:
                print(f"   - {sheet}")
        elif processed_sheets:
            print(f"\n📑 共处理 {len(processed_sheets)} 个工作表（列表过长，不逐一显示）")
        
        print(f"\n{'='*60}")
        
        # 弹出完成提示
        messagebox.showinfo("完成", 
                           f"汇总完成！\n\n"
                           f"输出文件：{os.path.basename(output_path)}\n"
                           f"处理文件：{len(files)} 个\n"
                           f"汇总数据：{total_rows} 行\n\n"
                           f"文件保存在：{output_dir}")
        
    else:
        messagebox.showwarning("警告", 
                              "未提取到任何有效数据！\n\n"
                              "请检查：\n"
                              "1. 工作表关键词是否正确\n"
                              "2. 指定的行范围内是否有数据\n"
                              "3. Excel文件是否为空")
        print("\n⚠ 未提取到任何有效数据，请检查参数设置")
    
    input("\n按回车键退出...")

if __name__ == "__main__":
    汇总Excel文件_按条件()