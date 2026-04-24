import warnings
# 抑制 openpyxl 相关的 UserWarning
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import os
import re
import sys
import gc
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl.utils import get_column_letter
import time

def progress_bar(current, total, bar_length=50):
    """显示进度条"""
    if total == 0:
        percent = 1.0
    else:
        percent = current / total
    filled_length = int(bar_length * percent)
    bar = '█' * filled_length + '-' * (bar_length - filled_length)
    # 添加 flush 确保在部分 IDE 或终端中实时显示
    print(f'\r[{bar}] {current}/{total} ({percent:.1%})', end='', flush=True)

def sanitize_sheet_name(name, max_len=31):
    """
    清理工作表名称，去除非法字符，并确保长度限制。
    注意：Excel sheet name 不能包含 \\ / ? * [ ] :
    """
    # 替换非法字符为下划线
    illegal_chars = r'[\/*?:\[\]]'
    sanitized = re.sub(illegal_chars, '_', str(name))
    # 去除首尾空格，因为 Excel 也不推荐
    sanitized = sanitized.strip()
    if len(sanitized) > max_len:
        sanitized = sanitized[:max_len]
    # 如果处理后为空，给予默认名
    if not sanitized:
        sanitized = "Sheet"
    return sanitized

def get_unique_sheet_name(writer, desired_name):
    """
    确保工作表名称在 workbook 中唯一。
    如果名称已存在，追加 _1, _2 等后缀。
    """
    final_name = desired_name
    counter = 1
    existing_sheets = writer.book.sheetnames
    while final_name in existing_sheets:
        # 计算基础名称和可能的后缀位置
        # 如果原始名称已经很长，需要截断以容纳后缀
        suffix = f"_{counter}"
        max_base_len = 31 - len(suffix)
        if len(desired_name) > max_base_len:
            base = desired_name[:max_base_len]
        else:
            base = desired_name
        final_name = f"{base}{suffix}"
        counter += 1
    return final_name

def auto_adjust_column_width(writer, sheet_name, df):
    try:
        workbook = writer.book
        worksheet = workbook[sheet_name]
        
        # 获取所有列索引
        cols = list(df.columns)
        if not cols:
            return

        for i, col in enumerate(cols):
            col_letter = get_column_letter(i + 1)
            
            # 计算标题长度
            header_len = len(str(col))
            
            # 计算数据最大长度
            max_data_len = 0
            if not df.empty:
                try:
                    # 转换为字符串并填充空值，避免 NaN 导致的问题
                    col_str = df[col].fillna('').astype(str)
                    if not col_str.empty:
                        # 使用 map(len) 获取每个元素的长度，然后取最大值
                        max_data_len = col_str.map(len).max()
                except Exception:
                    max_data_len = 0
            
            # 确定最终宽度：至少为标题长度，最多限制在 50，最小 5
            max_len = max(header_len, max_data_len)
            adjusted_width = min(max(max_len + 2, 5), 50)
            
            worksheet.column_dimensions[col_letter].width = adjusted_width
    except Exception as e:
        print(f"\n  警告：工作表 '{sheet_name}' 调整列宽时出错：{e}")

def merge_excel_files(file_paths, output_file):
    total_files = len(file_paths)
    if total_files == 0:
        print("没有文件需要处理。")
        return False

    print(f"开始合并 {total_files} 个Excel文件...")
    
    success = True
    # 使用 engine='openpyxl' 以支持 .xlsx 格式和样式调整
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for i, file_path in enumerate(file_paths, 1):
                # 更新进度条
                progress_bar(i, total_files)
                
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                try:
                    # sheet_name=None 读取所有 sheet，返回 OrderedDict
                    # dtype=str 可以防止某些列因混合类型被识别错误，但可能会影响性能，视情况而定
                    sheets = pd.read_excel(file_path, sheet_name=None)
                except Exception as e:
                    print(f"\n读取文件失败：{file_path}，错误：{e}")
                    continue
                
                if not sheets:
                    continue

                if len(sheets) == 1:
                    # 只有一个 sheet 的情况
                    original_sheet_name, df = list(sheets.items())[0]
                    # 使用文件名作为 sheet 名，通常更符合用户预期，或者使用原 sheet 名？
                    # 原代码逻辑是使用 base_name (文件名)。这里保持一致。
                    desired_name = sanitize_sheet_name(base_name)
                    final_name = get_unique_sheet_name(writer, desired_name)
                    
                    df.to_excel(writer, sheet_name=final_name, index=False)
                    auto_adjust_column_width(writer, final_name, df)
                else:
                    # 多个 sheet 的情况
                    for original_sheet_name, df in sheets.items():
                        # 组合 文件名_原Sheet名
                        combined_name = f"{base_name}_{original_sheet_name}"
                        desired_name = sanitize_sheet_name(combined_name)
                        final_name = get_unique_sheet_name(writer, desired_name)
                        
                        df.to_excel(writer, sheet_name=final_name, index=False)
                        auto_adjust_column_width(writer, final_name, df)
                
                # 每处理一个文件，可选地强制垃圾回收，防止内存爆炸（针对超大文件）
                # gc.collect() 
            
            # 完成进度条
            progress_bar(total_files, total_files)
            print()  # 换行
            
            # 显式保存并关闭，虽然 with 语句会做，但显式调用有助于调试
            # writer.save() # 在较新版本的 pandas 中，save() 已被弃用，close() 或直接退出 with 即可
            
    except Exception as e:
        print(f"\n错误 写入文件时发生严重错误：{e}")
        success = False
    finally:
        if success:
            print(f"\n成功 合并完成！")
            print(f"输出文件：{os.path.abspath(output_file)}")
            print(f"处理文件数：{total_files}")
            
            # 播放完成提示音（如果系统支持）
            try:
                import winsound
                winsound.Beep(800, 300)  # 频率800Hz，持续300ms
            except ImportError:
                pass # 非Windows系统忽略
            except Exception:
                pass # 其他错误忽略
        else:
            print("\n合并失败，请检查错误信息。")
            
    return success

if __name__ == "__main__":
    print("=" * 60)
    print("Excel文件合并工具")
    print("=" * 60)
    
    root = None
    try:
        # 初始化 Tkinter
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        
        # 立即弹出文件选择对话框
        file_paths = filedialog.askopenfilenames(
            title="请选择要合并的 Excel 文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not file_paths:
            print("未选择任何文件，程序退出。")
        else:
            output_file = filedialog.asksaveasfilename(
                title="保存合并后的文件",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not output_file:
                print("未指定输出文件，程序退出。")
            else:
                # 显示选择的文件
                print(f"\n已选择 {len(file_paths)} 个文件：")
                for file in file_paths:
                    print(f"  - {os.path.basename(file)}")
                print()
                
                # 执行合并
                start_time = time.time()
                # file_paths 是一个 tuple，转为 list 以便处理
                result = merge_excel_files(list(file_paths), output_file)
                end_time = time.time()
                
                if result:
                    print(f"耗时：{end_time - start_time:.2f}秒")
                
    except Exception as e:
        print(f"发生未知错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # 确保 Tkinter 窗口被正确销毁，防止进程挂起
        if root:
            root.destroy()
            # 在某些系统中，可能需要更新一下界面才能彻底关闭
            try:
                root.update()
            except:
                pass

    print("=" * 60)
    # 如果您不希望窗口自动关闭，可以取消下面一行的注释（会需要按回车）
    input("按回车键退出...")