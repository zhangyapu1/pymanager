import os
from tkinter import filedialog, Tk, simpledialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def 汇总多个Excel文件_指定工作表_openpyxl():
    """使用openpyxl汇总多个Excel文件（更精确的行列控制）"""
    
    # 隐藏tkinter主窗口
    root = Tk()
    root.withdraw()
    
    # 1. 输入参数
    start_row = simpledialog.askinteger("输入", "请输入起始行号：", initialvalue=48)
    if not start_row:
        print("操作已取消")
        input("按回车键退出...")
        return
    
    end_row = simpledialog.askinteger("输入", "请输入结束行号：", initialvalue=54)
    if not end_row:
        print("操作已取消")
        input("按回车键退出...")
        return
    
    target_sheet = simpledialog.askstring("输入", "请输入要汇总的工作表名称（输入 * 表示所有工作表）：", 
                                          initialvalue="*")
    if not target_sheet:
        print("操作已取消")
        input("按回车键退出...")
        return
    
    output_name = simpledialog.askstring("输入", "请输入输出文件名（不含扩展名）：", 
                                         initialvalue="汇总结果")
    if not output_name:
        print("操作已取消")
        input("按回车键退出...")
        return
    
    # 2. 选择多个Excel文件
    files = filedialog.askopenfilenames(
        title="请选择要汇总的Excel文件",
        filetypes=[("Excel文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]
    )
    
    if not files:
        print("未选择任何文件，程序退出")
        input("按回车键退出...")
        return
    
    print(f"已选择 {len(files)} 个文件，开始汇总...")
    
    # 3. 创建汇总工作簿
    wb_result = Workbook()
    ws_result = wb_result.active
    ws_result.title = "汇总结果"
    
    # 设置表头
    headers = ["来源文件", "工作表名称", "原始行号"]
    ws_result.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
    header_align = Alignment(horizontal="center")
    
    for cell in ws_result[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    current_row = 2
    total_rows = 0
    
    # 4. 处理每个文件
    for file_path in files:
        file_name = os.path.basename(file_path)
        print(f"正在处理: {file_name}")
        
        try:
            # 加载工作簿（只读模式，提高性能）
            wb_src = load_workbook(file_path, data_only=True, read_only=True)
            
            for sheet_name in wb_src.sheetnames:
                # 判断工作表名称是否符合条件
                if target_sheet != "*":
                    if target_sheet not in sheet_name:  # 模糊匹配
                        continue
                
                ws_src = wb_src[sheet_name]
                
                # 计算最大列数（在指定行范围内）
                max_col = 1
                for row in range(start_row, end_row + 1):
                    for col in range(1, ws_src.max_column + 1):
                        cell_value = ws_src.cell(row, col).value
                        if cell_value is not None and str(cell_value).strip() != "":
                            if col > max_col:
                                max_col = col
                
                # 提取数据
                row_count = 0
                for row in range(start_row, end_row + 1):
                    # 检查该行是否为空
                    is_empty = True
                    row_data = []
                    
                    for col in range(1, max_col + 1):
                        cell_value = ws_src.cell(row, col).value
                        if cell_value is not None and str(cell_value).strip() != "":
                            is_empty = False
                        row_data.append(cell_value if cell_value is not None else "")
                    
                    if not is_empty:
                        # 添加来源信息
                        row_with_source = [file_name, sheet_name, row] + row_data
                        ws_result.append(row_with_source)
                        current_row += 1
                        total_rows += 1
                        row_count += 1
                
                if row_count > 0:
                    print(f"  - 工作表 [{sheet_name}] 提取了 {row_count} 行数据")
            
            wb_src.close()
            
        except Exception as e:
            print(f"文件 [{file_name}] 处理失败: {str(e)}")
            continue
    
    # 5. 调整格式
    # 自动调整列宽
    for column in ws_result.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_result.column_dimensions[column_letter].width = adjusted_width
    
    # 冻结首行
    ws_result.freeze_panes = 'A2'
    
    # 6. 保存结果
    output_path = os.path.join(os.path.dirname(files[0]), f"{output_name}.xlsx")
    wb_result.save(output_path)
    
    print(f"\n汇总完成！")
    print(f"共处理 {len(files)} 个文件")
    print(f"共汇总 {total_rows} 行数据")
    print(f"结果已保存至: {output_path}")
    
    input("\n按回车键退出...")

if __name__ == "__main__":
    汇总多个Excel文件_指定工作表_openpyxl()